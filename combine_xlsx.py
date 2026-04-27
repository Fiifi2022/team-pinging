import sys, os, glob, re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

xlsx_dir = sys.argv[1]
out_file = sys.argv[2]
year_month = sys.argv[3]  # e.g. "2026-04"

BLUE = "1d6fc4"; BLUE_BG = "dbeafe"; GREEN_BG = "dcfce7"; AMBER_BG = "fef3c7"; GRAY_BG = "f4f4f5"

def hdr_font(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def hdr_fill(): return PatternFill("solid", fgColor=BLUE)
def cell_font(bold=False): return Font(name="Arial", size=10, bold=bold)
def thin(): s = Side(style="thin", color="D0D7DE"); return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center")
def autofit(ws, min_w=8, max_w=42):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def style_header(ws, row):
    for c in ws[row]:
        c.font = hdr_font()
        c.fill = hdr_fill()
        c.alignment = center()
        c.border = thin()

pattern = os.path.join(xlsx_dir, f"TeamPing_{year_month}-*.xlsx")
daily_files = sorted(glob.glob(pattern))

all_orders = []
daily_summary_rows = []

for fpath in daily_files:
    try:
        wb_day = load_workbook(fpath, data_only=True)

        # Orders sheet
        for sheet_name in wb_day.sheetnames:
            ws = wb_day[sheet_name]
            if sheet_name.startswith("Orders"):
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= 4:
                    headers = list(rows[2])
                    for row in rows[3:]:
                        if not any(row): 
                            continue
                        all_orders.append(dict(zip(headers, row)))
            elif sheet_name == "Daily Summaries":
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= 4:
                    headers = list(rows[2])
                    for row in rows[3:]:
                        if row and str(row[0]).startswith(year_month):
                            daily_summary_rows.append(dict(zip(headers, row)))
        wb_day.close()
    except Exception:
        pass

wb = Workbook()

# Monthly Summary
ws = wb.active
ws.title = "Monthly Summary"
ws.append([f"Team Ping — Monthly Reference Report — {year_month}"])
ws["A1"].font = Font(bold=True, name="Arial", size=16, color=BLUE)
ws.merge_cells("A1:D1")
ws.append([f"Compiled from {len(daily_files)} daily reports"])
ws["A2"].font = Font(name="Arial", size=11, color="52525b")
ws.append([])

completed = [o for o in all_orders if str(o.get("Status","")).lower() == "completed"]
pending = [o for o in all_orders if str(o.get("Status","")).lower() != "completed"]

def count(field, value):
    return sum(1 for o in all_orders if str(o.get(field,"")) == value)

def avg_mins():
    vals = []
    for o in all_orders:
        v = o.get("Time Taken (mins)")
        try:
            if v not in ("", None): vals.append(float(v))
        except Exception:
            pass
    return round(sum(vals)/len(vals)) if vals else ""

summary_rows = [
    ["Total orders", len(all_orders)],
    ["Completed", len(completed)],
    ["Pending", len(pending)],
    ["Medical", count("Type","Medical")],
    ["Vaccine", count("Type","Vaccine")],
    ["Emergency", count("Priority","Emergency")],
    ["Replenishment", count("Priority","Replenishment")],
    ["Scheduled", count("Priority","Scheduled")],
    ["Average completion mins", avg_mins()],
]
for row in summary_rows:
    ws.append(row)
for r in range(4, 4 + len(summary_rows)):
    for c in ws[r]:
        c.font = cell_font(c.column == 1)
        c.border = thin()
        c.alignment = center()
ws["A4"].fill = PatternFill("solid", fgColor=BLUE_BG)

ws.append([])
ws.append(["By Nest", "Total", "Completed", "Pending"])
style_header(ws, ws.max_row)
for nest in ["GH-1","GH-2","GH-3","GH-4","GH-5","GH-6"]:
    total = sum(1 for o in all_orders if o.get("Nest") == nest)
    done = sum(1 for o in completed if o.get("Nest") == nest)
    ws.append([nest, total, done, total-done])
    for c in ws[ws.max_row]:
        c.font = cell_font()
        c.border = thin()
        c.alignment = center()
autofit(ws)

# Daily Summaries sheet
ws2 = wb.create_sheet("Daily Summaries")
headers = ["Date","Total Orders","Completed","Pending","Medical","Vaccine","Emergency","Replenishment","Scheduled","Avg Completion (mins)","GH-1","GH-2","GH-3","GH-4","GH-5","GH-6"]
ws2.append([f"Daily Summaries — {year_month}"])
ws2["A1"].font = Font(bold=True, name="Arial", size=14, color=BLUE)
ws2.merge_cells("A1:P1")
ws2.append([])
ws2.append(headers)
style_header(ws2, 3)
for d in sorted(daily_summary_rows, key=lambda x: str(x.get("Date",""))):
    ws2.append([d.get(h,"") for h in headers])
    for c in ws2[ws2.max_row]:
        c.font = cell_font()
        c.border = thin()
        c.alignment = center()
autofit(ws2)

# All Orders sheet
ws3 = wb.create_sheet("All Orders")
order_headers = ["Order ID","Nest","Priority","Type","Submitted By","Time","Status","Completed By","Completed At","Time Taken (mins)","Notes","Link 1","Link 2"]
ws3.append([f"All Orders — {year_month}"])
ws3["A1"].font = Font(bold=True, name="Arial", size=14, color=BLUE)
ws3.merge_cells("A1:M1")
ws3.append([])
ws3.append(order_headers)
style_header(ws3, 3)
for o in all_orders:
    ws3.append([o.get(h,"") for h in order_headers])
    for c in ws3[ws3.max_row]:
        c.font = cell_font()
        c.border = thin()
        c.alignment = Alignment(wrap_text=True, vertical="top")
autofit(ws3)

# Included Files
ws4 = wb.create_sheet("Included Files")
ws4.append(["Included daily reports"])
ws4["A1"].font = Font(bold=True, name="Arial", size=14, color=BLUE)
for f in daily_files:
    ws4.append([os.path.basename(f)])
autofit(ws4)

wb.save(out_file)
print(f"Monthly report saved: {out_file} ({len(daily_files)} days, {len(all_orders)} orders)")
