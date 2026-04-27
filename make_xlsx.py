import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out_file = sys.argv[1]
date_str = sys.argv[2]
data = json.loads(sys.stdin.read())

orders = data.get("orders", [])
summaries = data.get("allSummaries", [])

BLUE  = "1d6fc4"; BLUE_BG = "dbeafe"
GREEN = "2d7d3a"; GREEN_BG = "dcfce7"
AMBER_BG = "fef3c7"; CORAL_BG = "fee2e2"
PURPLE_BG = "ede9fe"; TEAL_BG = "ccfbf1"
GRAY_BG = "f4f4f5"

def hdr_font(): return Font(bold=True, color="FFFFFF", name="Arial", size=11)
def hdr_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def cell_font(bold=False): return Font(name="Arial", size=10, bold=bold)
def thin_border():
    s = Side(style="thin", color="D0D7DE")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center")
def wrap(): return Alignment(wrap_text=True, vertical="top")

def style_header_row(ws, row, bg):
    for cell in ws[row]:
        cell.font = hdr_font()
        cell.fill = hdr_fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

wb = Workbook()

# ── Sheet 1: Today's orders ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = f"Orders {date_str}"

ws1.append([f"Team Ping — Daily Orders — {date_str}"])
ws1["A1"].font = Font(bold=True, name="Arial", size=14, color=BLUE)
ws1.merge_cells("A1:M1")
ws1.append([])

hdrs = ["Order ID","Nest","Priority","Type","Submitted By","Time","Status",
        "Completed By","Completed At","Time Taken (mins)","Notes","Link 1","Link 2"]
ws1.append(hdrs)
style_header_row(ws1, 3, BLUE)

PRIO_FILLS = {"Emergency": CORAL_BG, "Replenishment": AMBER_BG, "Scheduled": BLUE_BG}
TYPE_FILLS = {"Medical": PURPLE_BG, "Vaccine": TEAL_BG}

for o in orders:
    mins = ""
    if o.get("submitTimestamp") and o.get("doneTimestamp"):
        mins = round((o["doneTimestamp"] - o["submitTimestamp"]) / 60000)
    row = [
        o.get("orderID",""), o.get("nest",""), o.get("priority",""), o.get("type",""),
        o.get("from",""), o.get("time",""),
        "Completed" if o.get("done") else "Pending",
        o.get("doneBy",""), o.get("doneTime",""), mins,
        o.get("notes",""), o.get("link",""), o.get("link2","")
    ]
    ws1.append(row)
    r = ws1.max_row
    for c in ws1[r]:
        c.font = cell_font()
        c.border = thin_border()
        c.alignment = wrap()
    # Priority colour
    pf = PRIO_FILLS.get(o.get("priority",""))
    if pf: ws1.cell(r, 3).fill = PatternFill("solid", fgColor=pf)
    # Type colour
    tf = TYPE_FILLS.get(o.get("type",""))
    if tf: ws1.cell(r, 4).fill = PatternFill("solid", fgColor=tf)
    # Status colour
    if o.get("done"):
        ws1.cell(r, 7).fill = PatternFill("solid", fgColor=GREEN_BG)
        ws1.cell(r, 7).font = Font(name="Arial", size=10, color=GREEN, bold=True)
    else:
        ws1.cell(r, 7).fill = PatternFill("solid", fgColor=AMBER_BG)

ws1.freeze_panes = "A4"
ws1.row_dimensions[3].height = 18
autofit(ws1)

# ── Sheet 2: Daily summary history ────────────────────────────────────────────
ws2 = wb.create_sheet("Daily Summaries")
ws2.append(["Team Ping — Daily Summary History"])
ws2["A1"].font = Font(bold=True, name="Arial", size=14, color=BLUE)
ws2.merge_cells("A1:P1")
ws2.append([])

s_hdrs = ["Date","Total Orders","Completed","Pending","Medical","Vaccine",
          "Emergency","Replenishment","Scheduled","Avg Completion (mins)",
          "GH-1","GH-2","GH-3","GH-4","GH-5","GH-6"]
ws2.append(s_hdrs)
style_header_row(ws2, 3, BLUE)

for s in sorted(summaries, key=lambda x: x.get("date",""), reverse=True):
    bn = s.get("byNest", {})
    row = [
        s.get("date",""), s.get("archivedCount",0), s.get("completedCount",0),
        s.get("pendingCount",0), s.get("medicalCount",0), s.get("vaccineCount",0),
        s.get("emergencyCount",0), s.get("replenishmentCount",0), s.get("scheduledCount",0),
        s.get("avgCompletionMins",""),
        bn.get("GH-1",0), bn.get("GH-2",0), bn.get("GH-3",0),
        bn.get("GH-4",0), bn.get("GH-5",0), bn.get("GH-6",0)
    ]
    ws2.append(row)
    r = ws2.max_row
    for c in ws2[r]:
        c.font = cell_font()
        c.border = thin_border()
        c.alignment = center()
    if r % 2 == 0:
        for c in ws2[r]: c.fill = PatternFill("solid", fgColor=GRAY_BG)
    # Highlight today's row
    if s.get("date") == date_str:
        for c in ws2[r]: c.fill = PatternFill("solid", fgColor=BLUE_BG)

ws2.freeze_panes = "A4"
ws2.row_dimensions[3].height = 18
autofit(ws2)

wb.save(out_file)
print(f"Saved: {out_file}")
